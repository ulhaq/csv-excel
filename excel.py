import openpyxl
from csvalues import Csvalues

class Excel:
    def __init__(self, filename='iris_data.xlsx'):
        self.wb = openpyxl.load_workbook(filename)

    def getActiveSheet(self):
        return self.wb.active

    def getTotalRows(self):
            return self.getActiveSheet().max_row + 1

    def getTotalColumn(self):
        return self.getActiveSheet().max_column + 1

    def getColumn(self, pos = 1):
        list = []
        for i in range(1, self.getTotalRows()):
            cell = self.getActiveSheet().cell(row = i, column = pos)
            list.append(cell.value)
        return list

    def getRow(self, pos = 1):
        list = []
        for i in range(1, self.getTotalColumn()):
            cell = self.getActiveSheet().cell(row = pos, column = i)
            list.append(cell.value)
        return list

    def getContent(self):
        list = []

        for i in range(1, self.getTotalRows()):
            list.append(self.getRow(i))
        return list

    def xlToCsv(self, filename):
        c = Csvalues(filename)
        writer = c.getWriter()

        for i in range(1, self.getTotalRows()):
            writer.writerow(self.getRow(i))

        print('Excel data successfully converted to CSV')

        return
